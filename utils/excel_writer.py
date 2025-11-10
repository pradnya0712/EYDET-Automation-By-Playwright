import openpyxl
from datetime import datetime
import re
from openpyxl.utils.exceptions import IllegalCharacterError

# Regular expression to remove color/escape codes
_ANSI_ESCAPE = re.compile(r"\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])")


def clean_and_extract_status(value):
    """
    Clean the string for Excel and extract the final status part from an error message.
    Returns a safe string to write into Excel.
    """
    if not value:
        return ""
    # Ensure it's a string
    value = str(value).strip()
    # Remove terminal color codes
    value = _ANSI_ESCAPE.sub("", value)
    # Remove other non-printable ASCII characters
    value = "".join(ch for ch in value if ord(ch) >= 32 or ch in ("\t", "\n"))
    # Extract part after last colon, if any
    if ":" in value:
        value = value.split(":")[-1].strip()
    return value


def update_test_results_in_excel_sheet(output_file):
    wb = openpyxl.load_workbook("testdata/testcases.xlsx")
    sheet = wb["TestCase"]

    # Parse output file
    with open(output_file, "r") as f:
        output_lines = f.readlines()

    # Build result map
    result_map = {}
    for line in output_lines:
        parts = line.strip().split(" | ")
        id_and_status = parts[0].split()
        if len(id_and_status) < 2:
            continue
        test_case_id = id_and_status[0].strip()
        status = id_and_status[1].strip()
        error_msg = parts[1].strip() if len(parts) > 1 else ""
        result_map[test_case_id.upper()] = {
            "status": status,
            "actual_result": error_msg or status,
        }

    run_time_str = datetime.now().strftime("%d-%b-%Y %I:%M:%S %p")

    # Read headers safely
    headers = [
        sheet.cell(row=1, column=c).value or "" for c in range(1, sheet.max_column + 1)
    ]

    # Identify column sets
    col_sets = []
    max_suffix = 0
    idx = 0
    while idx < len(headers):
        header = headers[idx].strip().lower()
        if header.startswith("execution date"):
            exec_col = idx + 1
            actual_col = status_col = None
            match = re.search(r"execution date(\d*)", header)
            suffix = int(match.group(1)) if match and match.group(1) else 1
            max_suffix = max(max_suffix, suffix)
            if idx + 1 < len(headers) and headers[idx + 1].strip().lower().startswith(
                "actual result"
            ):
                actual_col = exec_col + 1
            if idx + 2 < len(headers) and headers[idx + 2].strip().lower().startswith(
                "status"
            ):
                status_col = exec_col + 2
            col_sets.append((exec_col, actual_col, status_col))
        idx += 1

    # Ensure at least one initial set exists
    if not col_sets:
        col_exec = sheet.max_column + 1
        col_actual = col_exec + 1
        col_status = col_exec + 2
        sheet.cell(row=1, column=col_exec, value="Execution Date1")
        sheet.cell(row=1, column=col_actual, value="Actual Result1")
        sheet.cell(row=1, column=col_status, value="Status1")
        col_sets.append((col_exec, col_actual, col_status))
        max_suffix = 1

    # Update each row
    for row in range(2, sheet.max_row + 1):
        test_case_id = sheet.cell(row=row, column=2).value
        if not test_case_id:
            continue
        key = test_case_id.strip().upper()
        if key not in result_map:
            continue
        result = result_map[key]

        updated = False
        for exec_col, actual_col, status_col in col_sets:
            if not actual_col or not status_col:
                continue
            last_status = sheet.cell(row=row, column=status_col).value
            last_status_clean = str(last_status).strip().lower() if last_status else ""
            if last_status_clean == "" or last_status_clean == "not executed":
                sheet.cell(row=row, column=exec_col, value=run_time_str)

                # Set actual_result blank if status is PASS
                actual_value = ""
                if result["status"].strip().upper() != "PASS":
                    actual_value = clean_and_extract_status(result["actual_result"])
                sheet.cell(row=row, column=actual_col, value=actual_value)

                sheet.cell(
                    row=row, column=status_col, value=result["status"].capitalize()
                )
                updated = True
                break

        if not updated:
            max_suffix += 1
            col_exec = sheet.max_column + 1
            col_actual = col_exec + 1
            col_status = col_exec + 2
            sheet.cell(row=1, column=col_exec, value=f"Execution Date{max_suffix}")
            sheet.cell(row=1, column=col_actual, value=f"Actual Result{max_suffix}")
            sheet.cell(row=1, column=col_status, value=f"Status{max_suffix}")
            sheet.cell(row=row, column=col_exec, value=run_time_str)

            actual_value = ""
            if result["status"].strip().upper() != "PASS":
                actual_value = clean_and_extract_status(result["actual_result"])
            sheet.cell(row=row, column=col_actual, value=actual_value)

            sheet.cell(row=row, column=col_status, value=result["status"].capitalize())
            col_sets.append((col_exec, col_actual, col_status))

    wb.save("testdata/testcases.xlsx")
    print(f"✅ Excel updated at {run_time_str}")
